"""
Schedule Parser for FieldConnect Intelligence Service.

Parses project schedules from various formats (CSV, Excel, P6 XML)
and converts them to a standard internal format.
"""

import csv
import re
from datetime import datetime, date
from pathlib import Path
from typing import Optional, BinaryIO
import logging

from ingestion.phase_mapper import map_activity_to_phase

logger = logging.getLogger(__name__)


def parse_date(date_str: str) -> Optional[date]:
    """
    Parse a date string in various formats.

    Supports:
    - ISO format: 2024-01-15
    - US format: 01/15/2024, 1/15/24
    - Written format: Jan 15, 2024
    """
    if not date_str or not isinstance(date_str, str):
        return None

    date_str = date_str.strip()
    if not date_str:
        return None

    # Try various date formats
    formats = [
        "%Y-%m-%d",       # ISO
        "%m/%d/%Y",       # US full year
        "%m/%d/%y",       # US short year
        "%d/%m/%Y",       # European full year
        "%d/%m/%y",       # European short year
        "%b %d, %Y",      # Jan 15, 2024
        "%B %d, %Y",      # January 15, 2024
        "%d-%b-%Y",       # 15-Jan-2024
        "%d-%b-%y",       # 15-Jan-24
        "%Y/%m/%d",       # Japanese style
    ]

    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue

    # Try parsing with dateutil as fallback
    try:
        from dateutil.parser import parse as dateutil_parse
        return dateutil_parse(date_str).date()
    except (ImportError, ValueError):
        pass

    logger.warning(f"Could not parse date: {date_str}")
    return None


def detect_csv_columns(header: list[str]) -> dict:
    """
    Auto-detect column mappings from CSV header.

    Returns dict mapping our standard fields to column indices.
    """
    mapping = {}
    header_lower = [h.lower().strip() for h in header]

    # Activity/Task name patterns
    activity_patterns = [
        "activity", "task", "activity name", "task name",
        "description", "activity description", "name"
    ]
    for i, h in enumerate(header_lower):
        for pattern in activity_patterns:
            if pattern in h:
                mapping["activity_name"] = i
                break
        if "activity_name" in mapping:
            break

    # Activity ID patterns
    id_patterns = ["activity id", "task id", "id", "wbs", "code"]
    for i, h in enumerate(header_lower):
        for pattern in id_patterns:
            if pattern in h and i not in mapping.values():
                mapping["activity_id"] = i
                break
        if "activity_id" in mapping:
            break

    # Start date patterns
    start_patterns = ["start", "start date", "planned start", "actual start", "begin"]
    for i, h in enumerate(header_lower):
        for pattern in start_patterns:
            if pattern in h:
                mapping["start_date"] = i
                break
        if "start_date" in mapping:
            break

    # End/Finish date patterns
    end_patterns = ["finish", "end", "end date", "finish date", "planned finish", "actual finish"]
    for i, h in enumerate(header_lower):
        for pattern in end_patterns:
            if pattern in h:
                mapping["end_date"] = i
                break
        if "end_date" in mapping:
            break

    # Duration patterns
    duration_patterns = ["duration", "days", "dur"]
    for i, h in enumerate(header_lower):
        for pattern in duration_patterns:
            if pattern in h:
                mapping["duration"] = i
                break
        if "duration" in mapping:
            break

    # Predecessor patterns
    pred_patterns = ["predecessor", "predecessors", "pred", "depends"]
    for i, h in enumerate(header_lower):
        for pattern in pred_patterns:
            if pattern in h:
                mapping["predecessors"] = i
                break
        if "predecessors" in mapping:
            break

    return mapping


def parse_csv_schedule(
    file_path: str,
    column_mapping: Optional[dict] = None
) -> list[dict]:
    """
    Parse a CSV schedule file.

    Args:
        file_path: Path to CSV file
        column_mapping: Optional dict mapping field names to column indices
                       If not provided, will auto-detect

    Returns:
        List of schedule entries with standardized fields
    """
    entries = []

    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, None)

        if not header:
            logger.warning(f"Empty CSV file: {file_path}")
            return entries

        # Auto-detect columns if mapping not provided
        if column_mapping is None:
            column_mapping = detect_csv_columns(header)

        if "activity_name" not in column_mapping:
            logger.error(f"Could not detect activity name column in {file_path}")
            return entries

        for row in reader:
            if not row or not any(row):
                continue

            entry = parse_schedule_row(row, column_mapping)
            if entry and entry.get("activity_name"):
                # Map activity to phase
                entry["phase"] = map_activity_to_phase(entry["activity_name"])
                entries.append(entry)

    logger.info(f"Parsed {len(entries)} activities from {file_path}")
    return entries


def parse_csv_schedule_from_content(
    content: str,
    column_mapping: Optional[dict] = None
) -> list[dict]:
    """
    Parse schedule from CSV content string.

    Args:
        content: CSV content as string
        column_mapping: Optional column mapping

    Returns:
        List of schedule entries
    """
    import io
    entries = []

    reader = csv.reader(io.StringIO(content))
    header = next(reader, None)

    if not header:
        return entries

    if column_mapping is None:
        column_mapping = detect_csv_columns(header)

    if "activity_name" not in column_mapping:
        logger.error("Could not detect activity name column")
        return entries

    for row in reader:
        if not row or not any(row):
            continue

        entry = parse_schedule_row(row, column_mapping)
        if entry and entry.get("activity_name"):
            entry["phase"] = map_activity_to_phase(entry["activity_name"])
            entries.append(entry)

    return entries


def parse_schedule_row(row: list[str], mapping: dict) -> Optional[dict]:
    """Parse a single row using the column mapping."""
    try:
        entry = {}

        # Activity name (required)
        if "activity_name" in mapping:
            idx = mapping["activity_name"]
            if idx < len(row):
                entry["activity_name"] = row[idx].strip()

        # Activity ID
        if "activity_id" in mapping:
            idx = mapping["activity_id"]
            if idx < len(row):
                entry["activity_id"] = row[idx].strip()

        # Start date
        if "start_date" in mapping:
            idx = mapping["start_date"]
            if idx < len(row):
                entry["start_date"] = parse_date(row[idx])

        # End date
        if "end_date" in mapping:
            idx = mapping["end_date"]
            if idx < len(row):
                entry["end_date"] = parse_date(row[idx])

        # Duration
        if "duration" in mapping:
            idx = mapping["duration"]
            if idx < len(row):
                try:
                    # Extract number from duration (e.g., "5d", "5 days", "5")
                    duration_str = row[idx].strip()
                    duration_match = re.match(r"(\d+)", duration_str)
                    if duration_match:
                        entry["duration_days"] = int(duration_match.group(1))
                except (ValueError, TypeError):
                    pass

        # Predecessors
        if "predecessors" in mapping:
            idx = mapping["predecessors"]
            if idx < len(row):
                entry["predecessors"] = row[idx].strip()

        return entry if entry.get("activity_name") else None

    except Exception as e:
        logger.warning(f"Error parsing schedule row: {e}")
        return None


def parse_excel_schedule(
    file_path: str,
    sheet_name: Optional[str] = None,
    column_mapping: Optional[dict] = None
) -> list[dict]:
    """
    Parse an Excel schedule file.

    Args:
        file_path: Path to Excel file (.xlsx or .xls)
        sheet_name: Name of sheet to parse (uses first sheet if not specified)
        column_mapping: Optional column mapping

    Returns:
        List of schedule entries
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        return []

    entries = []

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return entries

        # Get header and detect columns
        header = [str(cell) if cell else "" for cell in rows[0]]

        if column_mapping is None:
            column_mapping = detect_csv_columns(header)

        if "activity_name" not in column_mapping:
            logger.error(f"Could not detect activity name column in {file_path}")
            return entries

        # Process data rows
        for row in rows[1:]:
            if not row or not any(row):
                continue

            # Convert row to string list
            row_strings = [str(cell) if cell else "" for cell in row]

            entry = parse_schedule_row(row_strings, column_mapping)
            if entry and entry.get("activity_name"):
                # Handle Excel date objects directly
                if "start_date" in mapping and mapping["start_date"] < len(row):
                    cell_val = row[mapping["start_date"]]
                    if isinstance(cell_val, datetime):
                        entry["start_date"] = cell_val.date()
                    elif isinstance(cell_val, date):
                        entry["start_date"] = cell_val

                if "end_date" in mapping and mapping["end_date"] < len(row):
                    cell_val = row[mapping["end_date"]]
                    if isinstance(cell_val, datetime):
                        entry["end_date"] = cell_val.date()
                    elif isinstance(cell_val, date):
                        entry["end_date"] = cell_val

                entry["phase"] = map_activity_to_phase(entry["activity_name"])
                entries.append(entry)

        workbook.close()

    except Exception as e:
        logger.error(f"Error parsing Excel file {file_path}: {e}")

    logger.info(f"Parsed {len(entries)} activities from {file_path}")
    return entries


def parse_excel_schedule_from_bytes(
    content: bytes,
    sheet_name: Optional[str] = None,
    column_mapping: Optional[dict] = None
) -> list[dict]:
    """
    Parse schedule from Excel bytes content.

    Args:
        content: Excel file content as bytes
        sheet_name: Optional sheet name
        column_mapping: Optional column mapping

    Returns:
        List of schedule entries
    """
    try:
        import openpyxl
        import io
    except ImportError:
        logger.error("openpyxl not installed")
        return []

    entries = []

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True
        )

        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return entries

        header = [str(cell) if cell else "" for cell in rows[0]]

        if column_mapping is None:
            column_mapping = detect_csv_columns(header)

        if "activity_name" not in column_mapping:
            logger.error("Could not detect activity name column")
            return entries

        for row in rows[1:]:
            if not row or not any(row):
                continue

            row_strings = [str(cell) if cell else "" for cell in row]
            entry = parse_schedule_row(row_strings, column_mapping)

            if entry and entry.get("activity_name"):
                # Handle Excel date objects
                if "start_date" in column_mapping:
                    idx = column_mapping["start_date"]
                    if idx < len(row) and isinstance(row[idx], (datetime, date)):
                        entry["start_date"] = row[idx] if isinstance(row[idx], date) else row[idx].date()

                if "end_date" in column_mapping:
                    idx = column_mapping["end_date"]
                    if idx < len(row) and isinstance(row[idx], (datetime, date)):
                        entry["end_date"] = row[idx] if isinstance(row[idx], date) else row[idx].date()

                entry["phase"] = map_activity_to_phase(entry["activity_name"])
                entries.append(entry)

        workbook.close()

    except Exception as e:
        logger.error(f"Error parsing Excel content: {e}")

    return entries


def parse_schedule_file(file_path: str, **kwargs) -> list[dict]:
    """
    Parse a schedule file, auto-detecting the format.

    Args:
        file_path: Path to schedule file
        **kwargs: Additional arguments passed to format-specific parser

    Returns:
        List of schedule entries
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return parse_csv_schedule(file_path, **kwargs)
    elif suffix in [".xlsx", ".xls"]:
        return parse_excel_schedule(file_path, **kwargs)
    else:
        logger.error(f"Unsupported schedule format: {suffix}")
        return []


def get_schedule_date_range(schedule: list[dict]) -> Optional[tuple[date, date]]:
    """
    Get the overall date range of a schedule.

    Returns:
        Tuple of (earliest_start, latest_end) or None if no dates
    """
    starts = [e["start_date"] for e in schedule if e.get("start_date")]
    ends = [e["end_date"] for e in schedule if e.get("end_date")]

    if not starts and not ends:
        return None

    earliest = min(starts) if starts else None
    latest = max(ends) if ends else None

    if earliest and latest:
        return (earliest, latest)
    elif earliest:
        return (earliest, earliest)
    elif latest:
        return (latest, latest)

    return None


def get_phase_date_ranges(schedule: list[dict]) -> dict[str, tuple[date, date]]:
    """
    Get date ranges for each phase in the schedule.

    Returns:
        Dict mapping phase name to (start_date, end_date) tuple
    """
    phase_ranges: dict[str, tuple[list[date], list[date]]] = {}

    for entry in schedule:
        phase = entry.get("phase")
        if not phase:
            continue

        if phase not in phase_ranges:
            phase_ranges[phase] = ([], [])

        if entry.get("start_date"):
            phase_ranges[phase][0].append(entry["start_date"])
        if entry.get("end_date"):
            phase_ranges[phase][1].append(entry["end_date"])

    result = {}
    for phase, (starts, ends) in phase_ranges.items():
        if starts and ends:
            result[phase] = (min(starts), max(ends))

    return result


def validate_schedule(schedule: list[dict]) -> dict:
    """
    Validate a parsed schedule and return quality metrics.

    Returns:
        Dict with validation results and quality score
    """
    total = len(schedule)
    if total == 0:
        return {
            "valid": False,
            "error": "Empty schedule",
            "quality_score": 0.0
        }

    with_name = sum(1 for e in schedule if e.get("activity_name"))
    with_start = sum(1 for e in schedule if e.get("start_date"))
    with_end = sum(1 for e in schedule if e.get("end_date"))
    with_phase = sum(1 for e in schedule if e.get("phase"))

    # Calculate quality score (0-100)
    name_score = (with_name / total) * 25
    start_score = (with_start / total) * 25
    end_score = (with_end / total) * 25
    phase_score = (with_phase / total) * 25

    quality_score = name_score + start_score + end_score + phase_score

    # Date range check
    date_range = get_schedule_date_range(schedule)
    warnings = []

    if date_range:
        range_days = (date_range[1] - date_range[0]).days
        if range_days > 365 * 5:
            warnings.append("Schedule spans more than 5 years")
        if range_days < 7:
            warnings.append("Schedule spans less than a week")

    return {
        "valid": True,
        "total_activities": total,
        "activities_with_name": with_name,
        "activities_with_start": with_start,
        "activities_with_end": with_end,
        "activities_with_phase": with_phase,
        "phase_mapping_rate": round(with_phase / total * 100, 1) if total > 0 else 0,
        "date_range": date_range,
        "quality_score": round(quality_score, 1),
        "warnings": warnings
    }


# Sample schedule for testing when no real schedule is available
def create_sample_schedule(
    project_start: date,
    duration_months: int = 12
) -> list[dict]:
    """
    Create a sample construction schedule for testing.

    Args:
        project_start: Project start date
        duration_months: Total project duration in months

    Returns:
        Sample schedule with typical construction phases
    """
    from datetime import timedelta

    # Typical phase durations as percentage of total
    phase_durations = {
        "preconstruction": 0.08,
        "foundation": 0.10,
        "structure": 0.20,
        "envelope": 0.15,
        "mep_rough_in": 0.15,
        "interior_finishes": 0.18,
        "mep_trim_out": 0.08,
        "commissioning": 0.04,
        "closeout": 0.02
    }

    total_days = duration_months * 30
    schedule = []

    current_date = project_start
    for phase, pct in phase_durations.items():
        phase_days = int(total_days * pct)
        end_date = current_date + timedelta(days=phase_days)

        schedule.append({
            "activity_id": phase.upper(),
            "activity_name": phase.replace("_", " ").title(),
            "phase": phase,
            "start_date": current_date,
            "end_date": end_date,
            "duration_days": phase_days
        })

        # Overlap phases slightly (construction reality)
        current_date = current_date + timedelta(days=int(phase_days * 0.85))

    return schedule
